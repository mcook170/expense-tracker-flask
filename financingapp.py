from flask import Flask, render_template, request, redirect
from datetime import datetime
import openpyxl
import os

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Get form data
        description = request.form.get("description")
        expense_type = request.form.get("expense_type")
        amount = request.form.get("amount")
        note = request.form.get("note")
        date = datetime.now().strftime("%m/%d/%Y")

        # Load or create Excel file
        file_name = "expenses - app.xlsx"
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
        else:
            wb = openpyxl.Workbook()

        sheet = wb.active
        sheet.title = "Expenses - App"

        # Add headers if the file is new or empty
        if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
            sheet.append(["Date", "Type", "Description", "Amount", "Note"])

        # Add the expense data
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.0

        sheet.append([date, expense_type, description, amount, note])
        wb.save(file_name)
        print(f"Saved to: {os.path.abspath(file_name)}")
        print("Form submitted!")
        print(request.form)
        print(f"Saving to: {os.path.abspath(file_name)}")
        return redirect("/")

    return render_template("index.html")

# 🔥 This part starts the server
if __name__ == "__main__":
    app.run(debug=True)